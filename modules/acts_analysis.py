import csv
import os
import re
from pathlib import Path

from PyQt5.QtCore import QThread, pyqtSignal
import openpyxl


class ActsAnalysis(QThread):
    """Анализ Актов _xlsx_ с сохранением данных в _csv_ файл."""
    status_update = pyqtSignal(str)
    progress_update = pyqtSignal(int)

    def __init__(self, path_acts, path_output_csv_acts, stage):
        super().__init__()
        self.path_acts = path_acts
        self.path_output_csv_acts = path_output_csv_acts
        self.stage = stage

    def run(self):
        """Главная функция."""
        try:
            files = os.listdir(self.path_acts)
            if not files:
                self.status_update.emit('Нет файлов в папке для обработки')
            else:
                self.status_update.emit('Идет обработка актов, ожидайте...')
                step = 100 / len(files)
                progress = 0
                quantity = 0
                for file in files:
                    quantity += 1
                    progress += step
                    self.progress_update.emit(int(progress))
                    file_name = Path(file).stem
                    if Path(file).suffix == '.xlsx':
                        file_path = Path(self.path_acts, file)
                        wb = openpyxl.load_workbook(file_path)
                        wb_sheet = wb['Лист1']
                        self.file_processing(wb_sheet, file_name)
                self.progress_update.emit(100)
                self.status_update.emit(f'Готово. Обработано файлов: {quantity}')
        except Exception as e:
            print(e)

    def check_month(self, m1s, m1e, m2s='-', m2e='-', m3s='-', m3e='-'):
        """Распределение дат по своим месяцам."""

        stage = {
            '1 этап (01.08.2023 - 31.10.2023)': ['08', '09', '10'],
            '2 этап (01.11.2023 - 31.01.2024)': ['11', '12', '01'],
            '3 этап (01.02.2024 - 30.04.2024)': ['02', '03', '04'],
            '4 этап (01.05.2024 - 31.07.2024)': ['05', '06', '07'],
            '5 этап (01.08.2024 - 31.10.2024)': ['08', '09', '10'],
            '6 этап (01.11.2024 - 31.01.2025)': ['11', '12', '01'],
            '7 этап (01.02.2025 - 30.04.2025)': ['02', '03', '04'],
            '8 этап (01.05.2025 - 31.07.2025)': ['05', '06', '07'],
            '9 этап (01.08.2025 - 31.10.2025)': ['08', '09', '10'],
        }
        M1, M2, M3 = stage[self.stage.currentText()]
        months = (m1s, m1e, m2s, m2e, m3s, m3e)
        month_mapping = {M1: (0, 1), M2: (2, 3), M3: (4, 5)}
        result = ['-' for _ in range(6)]

        for i in range(0, len(months), 2):
            month = months[i]
            if month and month[3:5] in month_mapping:
                start_idx, end_idx = month_mapping[month[3:5]]
                result[start_idx] = month
                result[end_idx] = months[i + 1]

        return result

    @staticmethod
    def write_to_csv(data, file_name):
        """Создание CSV файла и внесение данных."""
        file_exists = os.path.isfile(file_name)
        with open(file_name, mode='a', newline='') as file:
            writer = csv.writer(file, delimiter=';')

            if not file_exists:
                writer.writerow(
                    ['ТТ', 'Тип', 'Наименование МО', 'Адрес',
                     'н1', 'к1', 'н2', 'к2', 'н3', 'к3', 'Подпись', 'Общее МО']
                )

            writer.writerow(data)

    def file_processing(self, sheet, file):
        """Просмотр файла."""
        title = None
        address = None
        csv_path = self.path_output_csv_acts

        for ir in range(1, sheet.max_row + 1):
            for ic in range(1, 3):
                obj = str(sheet.cell(ir, ic).value)
                if re.search(r'\*\d{3}\-\d{4}\*', obj):
                    point = obj
                    type_point = sheet[f'C{ir}'].value
                    month_1_start = sheet[f'D{ir}'].value
                    month_1_end = sheet[f'E{ir}'].value
                    month_2_start, month_2_end = "-", "-"
                    month_3_start, month_3_end = "-", "-"

                    if 'учреждение' in str(sheet[f'A{ir - 2}'].value):
                        title = sheet[f'A{ir - 2}'].value

                    if 'услуги:' in str(sheet[f'A{ir - 1}'].value):
                        address = sheet[f'A{ir - 1}'].value[23:]

                    if (not sheet.cell(ir + 1, 2).value
                            and sheet.cell(ir + 1, column=4).value):
                        month_2_start = sheet[f'D{ir + 1}'].value
                        month_2_end = sheet[f'E{ir + 1}'].value

                    if (not sheet.cell(ir + 2, 1).value
                            and sheet.cell(ir + 2, column=4).value
                            and month_2_start != "-"):
                        month_3_start = sheet[f'D{ir + 2}'].value
                        month_3_end = sheet[f'E{ir + 2}'].value

                    signature = sheet[f'J{sheet.max_row - 4}'].value
                    months = [month_1_start, month_1_end,
                              month_2_start, month_2_end,
                              month_3_start, month_3_end]
                    months = self.check_month(*months)
                    data_to_write = (point, type_point, title, address,
                                     *months, signature, file)

                    self.write_to_csv(data_to_write, csv_path)
