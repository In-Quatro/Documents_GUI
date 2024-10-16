import csv
import os
import re
from datetime import datetime, timedelta
from pathlib import Path

from PyQt5.QtCore import QThread, pyqtSignal
import openpyxl


class ActIncident(QThread):
    status_update = pyqtSignal(str)
    progress_update = pyqtSignal(int)

    def __init__(
            self,
            path_input_acts_incidents,
            path_csv_data_incidents,
            path_output_acts_incidents
    ):
        super().__init__()
        self.path_input_acts_incidents = path_input_acts_incidents
        self.path_csv_data_incidents = path_csv_data_incidents
        self.path_output_acts_incidents = path_output_acts_incidents
        self.incidents = []
        self.points = []

    def run(self):
        try:
            self.status_update.emit('Идет создание актов, ожидайте...')
            acts = os.listdir(self.path_input_acts_incidents)
            if not acts:
                self.status_update.emit(
                    'Нет актов в папке для внесения заявок')
            else:
                self.create_list_incident(self.path_csv_data_incidents)
                self.create_acts_with_incident(acts)
        except KeyError:
            self.status_update.emit(
                f'Ошибка: Проверьте файл с данными '
                f'"{self.path_csv_data_incidents.name}"')
        except Exception:
            self.update_status(f'Ошибка')

    @staticmethod
    def check_header(fieldnames):
        """Проверка наличия полей в _csv_ файле."""
        fieldnames_ = {'МО', 'ТТ', 'Номер заявки', 'Время назначения',
                       'Время в отложено', 'Время обработки', 'Время закрытия',
                       'Время ограничения', 'Коэффициент'}
        fieldnames = set(fieldnames)
        return fieldnames_ == fieldnames

    def create_list_incident(self, data):
        """Функция для создания листа с заявками из CSV файла."""
        with open(data, newline='') as csvfile:
            rows = csv.DictReader(csvfile, delimiter=';')
            for row in rows:
                self.incidents.append(row)
                self.points.append(row['ТТ'])

    def create_acts_with_incident(self, acts):
        """Функция для создания актов с заявками."""
        quantity = 0
        step = 100 / len(acts)
        progress = 0

        for act in acts:
            if Path(act).suffix == '.xlsx':
                file_path = Path(self.path_input_acts_incidents, act)
                wb = openpyxl.load_workbook(file_path)
                sheet = wb['Лист1']
                self.find_point(sheet)
                sheet.print_title_rows = '1:2'  # Сквозные строки

                # Сохранение файла в папку
                with Path(self.path_output_acts_incidents, act) as output_file:
                    wb.save(output_file)
                    quantity += 1
                    progress += step
                    self.progress_update.emit(int(progress))
        self.progress_update.emit(100)
        self.status_update.emit(f'Готово. Обработано файлов - {quantity}')

    @staticmethod
    def str_to_date(string, flag=False):
        """Функция преобразования даты."""
        if len(string) == 10:
            date_obj = datetime.strptime(string, '%d.%m.%Y')
            if flag:
                date_obj += timedelta(hours=23, minutes=59)
            return date_obj
        date_obj = datetime.strptime(string, '%d.%m.%Y %H:%M')
        return date_obj

    def fill_incident(self, sheet, idx, start, end, point):
        """Функция добавления заявок."""
        if self.incidents:
            for i, row in enumerate(self.incidents):
                if (
                        (row['ТТ'] == point and
                         start <= self.str_to_date(
                                    row['Время назначения']) <= end)
                ):
                    sheet[f'G{idx}'] = row['Номер заявки']
                    sheet[f'H{idx}'] = row['Время в отложено']
                    sheet[f'I{idx}'] = row['Время обработки']
                    sheet[f'J{idx}'] = row['Время назначения']
                    sheet[f'K{idx}'] = row['Время закрытия']
                    sheet[f'L{idx}'] = row['Время ограничения']
                    sheet[f'M{idx}'] = int(row['Коэффициент'])
                    self.incidents.pop(i)

    def check_month(self, sheet, point, idx):
        """Функция проверки наличия 2-го и 3-го месяца."""
        if not sheet[f'B{idx}'].value and sheet[f'D{idx}'].value:
            month_start = self.str_to_date(sheet[f'D{idx}'].value)
            month_end = self.str_to_date(sheet[f'E{idx}'].value, True)
            self.fill_incident(sheet, idx, month_start, month_end, point)

    def find_point(self, sheet):
        """Функция поиска технологических точек."""
        pattern_point = r'\*\d{3}\-\d{4}\*'

        for ir in range(1, sheet.max_row + 1):
            idx_2 = ir + 1
            idx_3 = ir + 2

            for ic in range(1, 3):
                obj = str(sheet.cell(ir, ic).value)
                point = obj

                if re.search(pattern_point, obj) and point in self.points:
                    # 1-й месяц
                    month_1_start = self.str_to_date(sheet[f'D{ir}'].value)
                    month_1_end = self.str_to_date(sheet[f'E{ir}'].value, True)
                    self.fill_incident(
                        sheet,
                        ir,
                        month_1_start,
                        month_1_end,
                        point
                    )

                    # 2-й месяц
                    self.check_month(sheet, point, idx_2)

                    # 3-й месяц
                    self.check_month(sheet, point, idx_3)
