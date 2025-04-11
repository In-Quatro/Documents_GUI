import csv
from pathlib import Path

from PyQt5.QtCore import QThread, pyqtSignal
import openpyxl

from modules.constants import (
    SIGNATURES_PARTIES, EXECUTIVE,
    SIGNATURE_MEDICAL_ORGANIZATION,
    MEDICAL_ORGANIZATION, STAMP_PLACE)
from modules.styles import (
    font, font_bold, border,
    alignment_1, alignment_2, alignment_3)
from modules.utils import get_new_file_name


class ActsCreate(QThread):
    """Создание Актов _xlsx_ из _csv_ файла с данными."""
    status_update = pyqtSignal(str)
    progress_update = pyqtSignal(int)

    def __init__(
            self,
            path_template_acts,
            path_csv_data_acts,
            path_output_acts,
            dialog
    ):
        super().__init__()
        self.path_template_acts = path_template_acts
        self.path_csv_data_acts = path_csv_data_acts
        self.path_output_acts = path_output_acts
        self.get_new_file_name = get_new_file_name
        self.dialog = dialog

    def run(self):
        try:
            self.post, self.fio = self.dialog
            self.status_update.emit('Идет создание актов, ожидайте...')
            self.file_processing(
                self.path_template_acts,
                self.path_output_acts,
                self.path_csv_data_acts
            )
        except Exception as e:
            print(e)

    @staticmethod
    def check_months(*args):
        """Обработка количества месяцев."""
        result = ['-'] * 6
        idx = 0
        for month in args:
            if month != '-':
                result[idx] = month
                idx += 1

        idx_month = str(idx // 2)  # Добавление количества месяцев
        result.append(idx_month)
        return result

    @staticmethod
    def fill_month_data(sheet, idx_row, i, month_start, month_end):
        """Заполнение строк датами, '-' и 1."""
        if month_start != '-':
            sheet[f'D{idx_row + i}'] = month_start
            sheet[f'E{idx_row + i}'] = month_end
            sheet[f'F{idx_row + i}'] = f'=E{idx_row + i}-D{idx_row + i}+1'

            for column in range(7, 13):
                sheet.cell(row=idx_row + i, column=column, value='−')
            sheet[f'M{idx_row + i}'] = 1

    def fill_signature(self, sheet, idx_row, signature):
        """Создание подписи в документе."""
        sheet[f'A{idx_row + 1}'] = SIGNATURES_PARTIES
        sheet[f'A{idx_row + 1}'].font = font
        sheet[f'A{idx_row + 1}'].alignment = alignment_1
        sheet.merge_cells(f'A{idx_row + 1}:M{idx_row + 1}')

        sheet[f'B{idx_row + 3}'] = EXECUTIVE
        sheet[f'B{idx_row + 3}'].font = font_bold

        sheet[f'J{idx_row + 3}'] = MEDICAL_ORGANIZATION
        sheet[f'J{idx_row + 3}'].font = font_bold

        sheet[f'B{idx_row + 5}'] = self.post
        sheet[f'B{idx_row + 5}'].font = font_bold
        sheet[f'B{idx_row + 5}'].alignment = alignment_2
        sheet.merge_cells(f'B{idx_row + 5}:D{idx_row + 5}')
        sheet.row_dimensions[idx_row + 5].height = 30

        sheet[f'J{idx_row + 5}'] = signature
        sheet[f'J{idx_row + 5}'].font = font_bold
        sheet[f'J{idx_row + 5}'].alignment = alignment_3
        sheet.merge_cells(f'J{idx_row + 5}:L{idx_row + 5}')

        sheet[f'B{idx_row + 8}'] = f'________________/{self.fio}/'
        sheet[f'B{idx_row + 8}'].font = font

        sheet[f'J{idx_row + 8}'] = SIGNATURE_MEDICAL_ORGANIZATION

        for i in ('B', 'J'):
            sheet[f'{i}{idx_row + 9}'] = STAMP_PLACE
            sheet[f'{i}{idx_row + 9}'].font = font

    @staticmethod
    def table_style(sheet):
        """Оформление таблицы."""
        for row in sheet.iter_rows(values_only=False):
            for cell in row:
                cell.border = border
                cell.font = font
                cell.alignment = alignment_1

                if isinstance(cell.value, str):
                    if 'государственное' in cell.value.lower():
                        cell.font = font_bold
                        cell.alignment = alignment_2
                    elif 'место' in cell.value.lower():
                        cell.alignment = alignment_2

        sheet.print_title_rows = '1:2'  # Сквозные строки для таблицы

    @staticmethod
    def get_step(file):
        with open(file, newline='') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            row_count = sum(1 for _ in reader)
            cnt = 100 / row_count
            return cnt

    def file_processing(self, template, folder, csvfile):
        """Создание Актов из CSV файла."""
        number = 0
        idx_row = 3
        current_address = None
        current_mo = None
        current_file = None
        wb = None
        current_signature = None
        step = self.get_step(csvfile)
        progress = 0
        quantity = 0
        self.progress_update.emit(progress)

        with open(csvfile, newline='') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            for row in reader:
                progress += step
                self.progress_update.emit(int(progress))
                point = row['ТТ']
                type_point = row['Тип']
                m_1_start = row['н1']
                m_1_end = row['к1']
                m_2_start = row['н2']
                m_2_end = row['к2']
                m_3_start = row['н3']
                m_3_end = row['к3']
                signature = row['Подпись']
                mo = row['Наименование МО']
                pref = 'Место оказания услуги: '
                address = row['Адрес']
                file = row['Общее МО']

                months = [m_1_start, m_1_end, m_2_start, m_2_end, m_3_start,
                          m_3_end]

                (m_1_start, m_1_end,
                 m_2_start, m_2_end,
                 m_3_start, m_3_end,
                 cnt_row) = self.check_months(*months)

                cnt_row = int(cnt_row)

                if current_file != file:
                    if wb is not None:
                        self.table_style(sheet)
                        self.fill_signature(sheet, idx_row, current_signature)
                        new_f_name = self.get_new_file_name(
                            f'{current_file}.xlsx',
                            folder
                        )
                        with Path(folder, new_f_name) as output_file:
                            wb.save(output_file)
                        quantity += 1

                    wb = openpyxl.load_workbook(template)
                    sheet = wb['Лист1']
                    current_file = file
                    idx_row = 3

                if current_address != address or current_mo != mo:
                    sheet[f'A{idx_row}'] = mo
                    sheet[f'A{idx_row + 1}'] = pref + address

                    sheet.merge_cells(f'A{idx_row}:M{idx_row}')
                    sheet.row_dimensions[idx_row].height = 30

                    sheet.merge_cells(f'A{idx_row + 1}:M{idx_row + 1}')
                    sheet.row_dimensions[idx_row + 1].height = 15

                    current_address = row['Адрес']
                    current_mo = row['Наименование МО']

                    idx_row += 2
                    number = 0

                sheet[f'A{idx_row}'] = number + 1
                sheet[f'B{idx_row}'] = point
                sheet[f'C{idx_row}'] = type_point

                self.fill_month_data(sheet, idx_row, 0, m_1_start, m_1_end)
                self.fill_month_data(sheet, idx_row, 1, m_2_start, m_2_end)
                self.fill_month_data(sheet, idx_row, 2, m_3_start, m_3_end)

                for i in ('A', 'B', 'C'):  # Объединение ячеек в таблице
                    sheet.merge_cells(
                        f'{i}{idx_row}:{i}{idx_row + cnt_row - 1}')

                idx_row += cnt_row
                number += 1
                current_signature = signature

            self.table_style(sheet)
            self.fill_signature(sheet, idx_row, current_signature)

            if current_file is not None:
                new_f_name = self.get_new_file_name(
                    f'{current_file}.xlsx', folder)
                with Path(folder, new_f_name) as output_file:
                    wb.save(output_file)
                quantity += 1
        self.progress_update.emit(100)
        self.status_update.emit(f'Готово. Создано файлов: {quantity}')
